import openpyxl
import random
import tkinter
import pyperclip
from tkinter.filedialog import askopenfilename

def main():
    # Fråga efter excel dokument
    filename = tkinter.filedialog.askopenfilename( filetypes =[('Excel files', '*.xlsx')])
    if filename == '':
        tkinter.messagebox.showerror("showerror", "No excel file selected") 
        return
    wb = openpyxl.load_workbook(filename)

    # Öppnar spellistan
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]

    # Skapar en array av alla spel
    games = []
    for i in range(2, worksheet.max_row+1):
        g = worksheet["{}{}".format("A", i)].value
        if g != None:
            games.append(g)
            # Lägger till extra kopior om det har lagts extra röster in i H kolumnen
            copies = worksheet["{}{}".format("H", i)].value
            if copies == None: copies = 0
            for j in range(int(copies)):
                games.append(g)
    print(games)
    print('Hur många spel vill du ha:')
    
    game_count = tkinter.simpledialog.askinteger('Antal spel','Hur många spel vill du ha')
    if game_count is None or game_count <= 0:
        #tkinter.messagebox.showerror("showerror", "Need atleast one game in game count\n") 
        # Sets to default 3 if nothing is selected
        game_count = 3

    selected_games = []
    # Lägg in antalet spel valda
    for i in range(int(game_count)):
        game = games[random.randrange(0, len(games))]
        selected_games.append(game)
        # Tar bort duplikationerna av valda spelet i listan
        games = [j for j in games if j != game]
        pass
    ret_str = ""
    for game in selected_games:
        ret_str += game + "\n"
    tkinter.messagebox.showinfo("Valda spel", ret_str)
    pyperclip.copy(ret_str)

    # x = input()

if __name__ == "__main__":
    main()